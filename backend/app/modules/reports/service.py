from __future__ import annotations

import csv
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from typing import Any, Callable

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session, joinedload

from app.models.branch import Branch
from app.models.compliance import ComplianceDocument
from app.models.farm import MortalityLog
from app.models.feed import FeedConsumption, FeedItem, FeedPurchase, FeedPurchaseItem
from app.models.finance import Expense, Income
from app.models.inventory import StockItem, StockMovement
from app.models.sales import Customer, Invoice, Order, OrderItem, Payment
from app.models.slaughter import SlaughterRecord
from app.services.pdf_branding import BRAND_DARK, BRAND_PRIMARY, draw_pdf_brand_header

from .schemas import ReportCatalogItem, ReportField, ReportFilter, ReportPreview, ReportRequest


def _field(key: str, label: str, default: bool = True) -> ReportField:
    return ReportField(key=key, label=label, default=default)


def _date_filters() -> list[ReportFilter]:
    return [
        ReportFilter(key="start_date", label="Start date", type="date"),
        ReportFilter(key="end_date", label="End date", type="date"),
        ReportFilter(key="search", label="Search", type="text"),
    ]


REPORT_CATALOG: dict[str, ReportCatalogItem] = {
    "sales-summary": ReportCatalogItem(
        key="sales-summary",
        title="Sales Summary",
        category="Sales Reports",
        description="Invoices, customers, payment status, amounts paid, and balances.",
        filters=_date_filters(),
        fields=[
            _field("invoice_number", "Invoice"),
            _field("date", "Date"),
            _field("customer", "Customer"),
            _field("status", "Status"),
            _field("total_amount", "Total"),
            _field("paid_amount", "Paid"),
            _field("balance", "Balance"),
        ],
    ),
    "payments": ReportCatalogItem(
        key="payments",
        title="Payments",
        category="Sales Reports",
        description="Payment collections by method, invoice, customer, and date.",
        filters=_date_filters(),
        fields=[
            _field("date", "Date"),
            _field("invoice_number", "Invoice"),
            _field("customer", "Customer"),
            _field("method", "Method"),
            _field("amount", "Amount"),
            _field("reference", "Reference"),
        ],
    ),
    "inventory-stock": ReportCatalogItem(
        key="inventory-stock",
        title="Inventory Stock",
        category="Inventory Reports",
        description="Current stock balances, reorder levels, costs, and stock value.",
        filters=[ReportFilter(key="search", label="Search", type="text")],
        fields=[
            _field("item_name", "Item"),
            _field("category", "Category"),
            _field("quantity", "Quantity"),
            _field("unit", "Unit"),
            _field("reorder_level", "Reorder"),
            _field("average_cost", "Avg cost"),
            _field("stock_value", "Stock value"),
            _field("status", "Status"),
        ],
    ),
    "stock-movements": ReportCatalogItem(
        key="stock-movements",
        title="Stock Movements",
        category="Inventory Reports",
        description="Incoming, outgoing, and adjustment movement history.",
        filters=_date_filters(),
        fields=[
            _field("date", "Date"),
            _field("item_name", "Item"),
            _field("movement_type", "Type"),
            _field("quantity", "Quantity"),
            _field("previous_quantity", "Previous"),
            _field("new_quantity", "New"),
            _field("reference", "Reference"),
        ],
    ),
    "feed-stock": ReportCatalogItem(
        key="feed-stock",
        title="Feed Stock",
        category="Feed Reports",
        description="Feed item balances and low-stock status.",
        filters=[ReportFilter(key="search", label="Search", type="text")],
        fields=[
            _field("item_name", "Feed item"),
            _field("category", "Category"),
            _field("current_stock", "Current stock"),
            _field("unit", "Unit"),
            _field("reorder_threshold", "Reorder threshold"),
            _field("status", "Status"),
        ],
    ),
    "feed-purchases": ReportCatalogItem(
        key="feed-purchases",
        title="Feed Purchases",
        category="Feed Reports",
        description="Feed purchases by supplier, invoice, line item, and amount.",
        filters=_date_filters(),
        fields=[
            _field("date", "Date"),
            _field("supplier", "Supplier"),
            _field("invoice_number", "Invoice"),
            _field("feed_item", "Feed item"),
            _field("quantity", "Quantity"),
            _field("unit_price", "Unit price"),
            _field("total_price", "Total"),
        ],
    ),
    "feed-consumption": ReportCatalogItem(
        key="feed-consumption",
        title="Feed Consumption",
        category="Feed Reports",
        description="Feed usage by date, batch, and feed item.",
        filters=_date_filters(),
        fields=[
            _field("date", "Date"),
            _field("batch", "Batch"),
            _field("feed_item", "Feed item"),
            _field("quantity", "Quantity"),
            _field("notes", "Notes", False),
        ],
    ),
    "profit-loss": ReportCatalogItem(
        key="profit-loss",
        title="Profit and Loss",
        category="Finance Reports",
        description="Income and expense lines with net contribution totals.",
        filters=_date_filters(),
        fields=[
            _field("date", "Date"),
            _field("type", "Type"),
            _field("category", "Category"),
            _field("description", "Description"),
            _field("reference", "Reference", False),
            _field("amount", "Amount"),
        ],
    ),
    "compliance-expiring": ReportCatalogItem(
        key="compliance-expiring",
        title="Expiring Compliance",
        category="Compliance Reports",
        description="Compliance documents that are expired, due soon, or need renewal.",
        filters=[ReportFilter(key="end_date", label="Due before", type="date"), ReportFilter(key="search", label="Search", type="text")],
        fields=[
            _field("title", "Document"),
            _field("type", "Type"),
            _field("reference_number", "Reference"),
            _field("expiry_date", "Expiry"),
            _field("days_left", "Days left"),
            _field("status", "Status"),
            _field("responsible_person", "Responsible"),
        ],
    ),
    "debtors": ReportCatalogItem(
        key="debtors",
        title="Debtors / Outstanding Balances",
        category="Sales Reports",
        description="All customers with unpaid or partially paid invoices, showing outstanding balances and due dates.",
        filters=[
            ReportFilter(key="start_date", label="Due from", type="date"),
            ReportFilter(key="end_date", label="Due to", type="date"),
            ReportFilter(key="search", label="Customer / Invoice", type="text"),
        ],
        fields=[
            _field("customer", "Customer"),
            _field("invoice_number", "Invoice"),
            _field("issue_date", "Issued"),
            _field("due_date", "Due date"),
            _field("total_amount", "Invoice total"),
            _field("paid_amount", "Paid"),
            _field("balance", "Outstanding"),
            _field("days_overdue", "Days overdue"),
            _field("status", "Status"),
        ],
    ),
    "credit-sales": ReportCatalogItem(
        key="credit-sales",
        title="Credit Sales",
        category="Sales Reports",
        description="All credit and partial-payment sales with outstanding amounts per customer.",
        filters=_date_filters(),
        fields=[
            _field("invoice_number", "Invoice"),
            _field("date", "Date"),
            _field("customer", "Customer"),
            _field("phone", "Phone", False),
            _field("total_amount", "Total"),
            _field("paid_amount", "Paid"),
            _field("balance", "Balance"),
            _field("due_date", "Due date"),
            _field("status", "Status"),
        ],
    ),
    # ── Accounting Reports ──────────────────────────────────────────
    "balance-sheet": ReportCatalogItem(
        key="balance-sheet",
        title="Balance Sheet",
        category="Accounting Reports",
        description="Assets, liabilities and equity at a given date.",
        filters=[
            ReportFilter(key="as_of_date", label="As of date", type="date"),
            ReportFilter(key="branch_id",  label="Branch (optional)", type="text"),
        ],
        fields=[
            _field("section",       "Section"),
            _field("account_code",  "Code"),
            _field("account_name",  "Account"),
            _field("amount",        "Amount (UGX)"),
        ],
    ),
    "general-ledger": ReportCatalogItem(
        key="general-ledger",
        title="General Ledger",
        category="Accounting Reports",
        description="Full transaction history with running balance for any account.",
        filters=[
            ReportFilter(key="account_id", label="Account ID", type="text"),
            ReportFilter(key="from_date",  label="From date",  type="date"),
            ReportFilter(key="end_date",   label="To date",    type="date"),
            ReportFilter(key="branch_id",  label="Branch",     type="text"),
        ],
        fields=[
            _field("date",        "Date"),
            _field("reference",   "Reference"),
            _field("description", "Description"),
            _field("debit",       "Debit (UGX)"),
            _field("credit",      "Credit (UGX)"),
            _field("balance",     "Balance (UGX)"),
        ],
    ),
    "cashbook": ReportCatalogItem(
        key="cashbook",
        title="Cashbook",
        category="Accounting Reports",
        description="Cash inflows and outflows for a cash or bank account.",
        filters=[
            ReportFilter(key="account_id", label="Account ID", type="text"),
            ReportFilter(key="from_date",  label="From date",  type="date"),
            ReportFilter(key="end_date",   label="To date",    type="date"),
        ],
        fields=[
            _field("date",        "Date"),
            _field("description", "Description"),
            _field("receipts",    "Receipts (IN)"),
            _field("payments",    "Payments (OUT)"),
            _field("balance",     "Balance (UGX)"),
        ],
    ),
    "cash-flow": ReportCatalogItem(
        key="cash-flow",
        title="Cash Flow Statement",
        category="Accounting Reports",
        description="Cash flow statement showing operating, investing, and financing activities.",
        filters=[
            ReportFilter(key="from_date",  label="From date",  type="date"),
            ReportFilter(key="end_date",   label="To date",    type="date"),
            ReportFilter(key="branch_id",  label="Branch",     type="text"),
        ],
        fields=[
            _field("section",     "Section"),
            _field("category",    "Category"),
            _field("amount",      "Amount (UGX)"),
        ],
    ),
    # ── Sales / Procurement ──────────────────────────────────────────
    "ar-aging": ReportCatalogItem(
        key="ar-aging",
        title="Accounts Receivable Aging",
        category="Sales Reports",
        description="Outstanding customer balances aged by days overdue.",
        filters=_date_filters(),
        fields=[
            _field("customer",   "Customer"),
            _field("current",    "Current (0-30 days)"),
            _field("days_31_60", "31-60 days"),
            _field("days_61_90", "61-90 days"),
            _field("over_90",    "Over 90 days"),
            _field("total",      "Total Outstanding"),
        ],
    ),
    "ap-aging": ReportCatalogItem(
        key="ap-aging",
        title="Accounts Payable Aging",
        category="Procurement Reports",
        description="Outstanding supplier invoice balances aged by days overdue.",
        filters=_date_filters(),
        fields=[
            _field("supplier",   "Supplier"),
            _field("current",    "Current (0-30 days)"),
            _field("days_31_60", "31-60 days"),
            _field("days_61_90", "61-90 days"),
            _field("over_90",    "Over 90 days"),
            _field("total",      "Total Outstanding"),
        ],
    ),
    # ── Farm Reports ─────────────────────────────────────────────────
    "cost-per-bird": ReportCatalogItem(
        key="cost-per-bird",
        title="Cost Per Bird Analysis",
        category="Farm Reports",
        description="Production cost breakdown per batch: feed, labour, overhead, cost/kg.",
        filters=_date_filters(),
        fields=[
            _field("batch",          "Batch"),
            _field("slaughter_date", "Date"),
            _field("birds_in",       "Birds"),
            _field("dressed_weight", "Dressed Wt (kg)"),
            _field("feed_cost",      "Feed Cost"),
            _field("labour_cost",    "Labour"),
            _field("overhead_cost",  "Overhead"),
            _field("total_cost",     "Total Cost"),
            _field("cost_per_bird",  "Cost/Bird"),
            _field("cost_per_kg",    "Cost/kg"),
        ],
    ),
    "mortality-analysis": ReportCatalogItem(
        key="mortality-analysis",
        title="Mortality Analysis",
        category="Farm Reports",
        description="Bird mortality by batch, cause and period.",
        filters=_date_filters(),
        fields=[
            _field("date",      "Date"),
            _field("batch",     "Batch"),
            _field("cause",     "Cause"),
            _field("count",     "Dead Count"),
            _field("pct_flock", "% of Flock"),
        ],
    ),
    "slaughter-yield": ReportCatalogItem(
        key="slaughter-yield",
        title="Slaughter Yield Report",
        category="Farm Reports",
        description="Slaughter batch yield percentage, carcass weight and by-products.",
        filters=_date_filters(),
        fields=[
            _field("batch",          "Batch"),
            _field("slaughter_date", "Date"),
            _field("birds_in",       "Birds In"),
            _field("live_weight",    "Live Wt (kg)"),
            _field("dressed_weight", "Dressed Wt (kg)"),
            _field("yield_pct",      "Yield %"),
            _field("condemned",      "Condemned"),
            _field("cost_per_kg",    "Cost/kg"),
        ],
    ),
    # ── Management Reports ───────────────────────────────────────────
    "branch-performance": ReportCatalogItem(
        key="branch-performance",
        title="Branch Performance Comparison",
        category="Management Reports",
        description="Revenue, COGS, gross margin and net income per branch side by side.",
        filters=_date_filters(),
        fields=[
            _field("branch",       "Branch"),
            _field("revenue",      "Revenue"),
            _field("cogs",         "COGS"),
            _field("gross_margin", "Gross Margin"),
            _field("expenses",     "Operating Exp."),
            _field("net_income",   "Net Income"),
            _field("margin_pct",   "Margin %"),
        ],
    ),
}


def _selected_fields(report: ReportCatalogItem, request: ReportRequest) -> list[str]:
    allowed = {field.key for field in report.fields}
    selected = [field for field in request.selected_fields if field in allowed]
    if selected:
        return selected
    return [field.key for field in report.fields if field.default]


def _date_range(request: ReportRequest, default_days: int = 30) -> tuple[date, date]:
    today = date.today()
    end = request.end_date or today
    start = request.start_date or (end - timedelta(days=default_days))
    if start > end:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Start date cannot be after end date")
    return start, end


def _label(value) -> str:
    raw = getattr(value, "value", value)
    return str(raw or "").replace("_", " ").title()


def _contains_search(*columns, search: str | None = None):
    if not search:
        return None
    pattern = f"%{search.strip()}%"
    return or_(*[column.ilike(pattern) for column in columns])


class ReportsService:
    def catalog(self) -> list[ReportCatalogItem]:
        return list(REPORT_CATALOG.values())

    def _report(self, report_key: str) -> ReportCatalogItem:
        report = REPORT_CATALOG.get(report_key)
        if not report:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found")
        return report

    def preview(self, db: Session, report_key: str, request: ReportRequest) -> ReportPreview:
        report = self._report(report_key)
        rows, totals = self._rows_for_report(db, report_key, request)
        selected = _selected_fields(report, request)
        filtered_rows = [{field: row.get(field, "") for field in selected} for row in rows[: request.limit]]
        return ReportPreview(
            report=report,
            selected_fields=selected,
            filters_applied={
                "start_date": request.start_date.isoformat() if request.start_date else None,
                "end_date": request.end_date.isoformat() if request.end_date else None,
                "search": request.search,
            },
            rows=filtered_rows,
            totals=totals,
            row_count=len(rows),
        )

    def export(self, db: Session, report_key: str, request: ReportRequest, export_format: str) -> tuple[bytes, str, str]:
        preview = self.preview(db, report_key, request)
        filename_base = f"farmexa-{report_key}-{datetime.now().strftime('%Y%m%d%H%M')}"
        if export_format == "csv":
            return self._csv(preview), f"{filename_base}.csv", "text/csv"
        if export_format == "xlsx":
            return self._xlsx(preview), f"{filename_base}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if export_format == "pdf":
            return self._pdf(preview), f"{filename_base}.pdf", "application/pdf"
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported export format")

    def _rows_for_report(self, db: Session, report_key: str, request: ReportRequest) -> tuple[list[dict[str, object]], dict[str, float | int | str]]:
        handlers: dict[str, Callable[[Session, ReportRequest], tuple[list[dict[str, object]], dict[str, float | int | str]]]] = {
            "sales-summary": self._sales_summary,
            "payments": self._payments,
            "inventory-stock": self._inventory_stock,
            "stock-movements": self._stock_movements,
            "feed-stock": self._feed_stock,
            "feed-purchases": self._feed_purchases,
            "feed-consumption": self._feed_consumption,
            "profit-loss": self._profit_loss,
            "compliance-expiring": self._compliance_expiring,
            "debtors": self._debtors,
            "credit-sales": self._credit_sales,
            # Accounting reports
            "balance-sheet": self._balance_sheet,
            "general-ledger": self._general_ledger,
            "cashbook": self._cashbook_report,
            "cash-flow": self._cash_flow_report,
            # Sales / procurement
            "ar-aging": self._ar_aging,
            "ap-aging": self._ap_aging,
            # Farm reports
            "cost-per-bird": self._cost_per_bird,
            "mortality-analysis": self._mortality_analysis,
            "slaughter-yield": self._slaughter_yield,
            # Management
            "branch-performance": self._branch_performance,
        }
        if report_key not in handlers:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"No data handler for report '{report_key}'")
        return handlers[report_key](db, request)

    def _sales_summary(self, db: Session, request: ReportRequest):
        start, end = _date_range(request)
        query = (
            db.query(Invoice)
            .options(joinedload(Invoice.customer))
            .filter(Invoice.issue_date >= start, Invoice.issue_date <= end)
        )
        search_filter = _contains_search(Invoice.invoice_number, search=request.search)
        if search_filter is not None:
            query = query.filter(search_filter)
        invoices = query.order_by(Invoice.issue_date.desc(), Invoice.id.desc()).all()
        rows = [
            {
                "invoice_number": invoice.invoice_number,
                "date": invoice.issue_date.isoformat(),
                "customer": invoice.customer.name if invoice.customer else "Walk-in",
                "status": _label(invoice.status),
                "total_amount": float(invoice.total_amount or 0),
                "paid_amount": float(invoice.paid_amount or 0),
                "balance": max(float(invoice.total_amount or 0) - float(invoice.paid_amount or 0), 0),
            }
            for invoice in invoices
        ]
        return rows, {"total": sum(row["total_amount"] for row in rows), "paid": sum(row["paid_amount"] for row in rows), "balance": sum(row["balance"] for row in rows)}

    def _payments(self, db: Session, request: ReportRequest):
        start, end = _date_range(request)
        query = (
            db.query(Payment)
            .join(Invoice)
            .options(joinedload(Payment.invoice).joinedload(Invoice.customer))
            .filter(Payment.payment_date >= start, Payment.payment_date <= end)
        )
        search_filter = _contains_search(Invoice.invoice_number, search=request.search)
        if search_filter is not None:
            query = query.filter(search_filter)
        payments = query.order_by(Payment.payment_date.desc(), Payment.id.desc()).all()
        rows = [
            {
                "date": payment.payment_date.isoformat(),
                "invoice_number": payment.invoice.invoice_number,
                "customer": payment.invoice.customer.name if payment.invoice.customer else "Walk-in",
                "method": _label(payment.payment_method),
                "amount": float(payment.amount or 0),
                "reference": payment.reference or "",
            }
            for payment in payments
        ]
        return rows, {"total_collected": sum(row["amount"] for row in rows), "payments": len(rows)}

    def _inventory_stock(self, db: Session, request: ReportRequest):
        query = db.query(StockItem)
        search_filter = _contains_search(StockItem.name, StockItem.sku, search=request.search)
        if search_filter is not None:
            query = query.filter(search_filter)
        items = query.order_by(StockItem.name.asc()).all()
        rows = [
            {
                "item_name": item.name,
                "category": _label(item.category),
                "quantity": float(item.current_quantity or 0),
                "unit": item.unit_of_measure,
                "reorder_level": float(item.reorder_level or 0),
                "average_cost": float(item.average_cost or 0),
                "stock_value": float(item.current_quantity or 0) * float(item.average_cost or 0),
                "status": "Low stock" if float(item.current_quantity or 0) <= float(item.reorder_level or 0) else "Active",
            }
            for item in items
        ]
        return rows, {"items": len(rows), "low_stock": len([row for row in rows if row["status"] == "Low stock"]), "stock_value": sum(row["stock_value"] for row in rows)}

    def _stock_movements(self, db: Session, request: ReportRequest):
        start, end = _date_range(request)
        query = (
            db.query(StockMovement)
            .options(joinedload(StockMovement.item))
            .filter(func.date(StockMovement.created_at) >= start, func.date(StockMovement.created_at) <= end)
        )
        movements = query.order_by(StockMovement.created_at.desc()).all()
        rows = [
            {
                "date": movement.created_at.date().isoformat(),
                "item_name": movement.item.name if movement.item else f"Item #{movement.item_id}",
                "movement_type": _label(movement.movement_type),
                "quantity": float(movement.quantity or 0),
                "previous_quantity": float(movement.previous_quantity or 0),
                "new_quantity": float(movement.new_quantity or 0),
                "reference": f"{movement.reference_type or ''} {movement.reference_id or ''}".strip(),
            }
            for movement in movements
        ]
        return rows, {"movements": len(rows)}

    def _feed_stock(self, db: Session, request: ReportRequest):
        query = db.query(FeedItem).options(joinedload(FeedItem.category), joinedload(FeedItem.stock_item))
        search_filter = _contains_search(FeedItem.name, search=request.search)
        if search_filter is not None:
            query = query.filter(search_filter)
        items = query.order_by(FeedItem.name.asc()).all()
        rows = []
        for item in items:
            linked_stock = item.stock_item
            quantity = float(linked_stock.current_quantity or 0) if linked_stock else 0.0
            status = "Unlinked inventory item" if linked_stock is None else (
                "Low stock" if quantity <= float(item.reorder_threshold or 0) else "Active"
            )
            rows.append(
                {
                    "item_name": item.name,
                    "category": item.category.name if item.category else "Uncategorized",
                    "current_stock": quantity,
                    "unit": linked_stock.unit_of_measure if linked_stock else item.unit,
                    "reorder_threshold": float(item.reorder_threshold or 0),
                    "status": status,
                }
            )
        return rows, {"items": len(rows), "low_stock": len([row for row in rows if row["status"] == "Low stock"])}

    def _feed_purchases(self, db: Session, request: ReportRequest):
        start, end = _date_range(request)
        query = (
            db.query(FeedPurchaseItem)
            .join(FeedPurchase)
            .options(joinedload(FeedPurchaseItem.purchase).joinedload(FeedPurchase.supplier), joinedload(FeedPurchaseItem.feed_item))
            .filter(FeedPurchase.purchase_date >= start, FeedPurchase.purchase_date <= end)
        )
        records = query.order_by(FeedPurchase.purchase_date.desc()).all()
        rows = [
            {
                "date": item.purchase.purchase_date.isoformat(),
                "supplier": item.purchase.supplier.name if item.purchase.supplier else "",
                "invoice_number": item.purchase.invoice_number or "",
                "feed_item": item.feed_item.name if item.feed_item else "",
                "quantity": float(item.quantity or 0),
                "unit_price": float(item.unit_price or 0),
                "total_price": float(item.total_price or 0),
            }
            for item in records
        ]
        return rows, {"total": sum(row["total_price"] for row in rows), "lines": len(rows)}

    def _feed_consumption(self, db: Session, request: ReportRequest):
        start, end = _date_range(request)
        query = (
            db.query(FeedConsumption)
            .options(joinedload(FeedConsumption.batch), joinedload(FeedConsumption.feed_item))
            .filter(FeedConsumption.record_date >= start, FeedConsumption.record_date <= end)
        )
        records = query.order_by(FeedConsumption.record_date.desc()).all()
        rows = [
            {
                "date": record.record_date.isoformat(),
                "batch": record.batch.batch_number if record.batch else "",
                "feed_item": record.feed_item.name if record.feed_item else "",
                "quantity": float(record.quantity or 0),
                "notes": record.notes or "",
            }
            for record in records
        ]
        return rows, {"quantity": sum(row["quantity"] for row in rows), "records": len(rows)}

    def _profit_loss(self, db: Session, request: ReportRequest):
        start, end = _date_range(request)
        expenses = (
            db.query(Expense)
            .options(joinedload(Expense.category))
            .filter(Expense.expense_date >= start, Expense.expense_date <= end)
            .all()
        )
        incomes = (
            db.query(Income)
            .options(joinedload(Income.category))
            .filter(Income.income_date >= start, Income.income_date <= end)
            .all()
        )
        rows = [
            {
                "date": income.income_date.isoformat(),
                "type": "Income",
                "category": income.category.name if income.category else "Income",
                "description": income.description or "",
                "reference": income.reference or "",
                "amount": float(income.amount or 0),
            }
            for income in incomes
        ] + [
            {
                "date": expense.expense_date.isoformat(),
                "type": "Expense",
                "category": expense.category.name if expense.category else "Expense",
                "description": expense.description or "",
                "reference": expense.reference or "",
                "amount": -float(expense.amount or 0),
            }
            for expense in expenses
        ]
        rows.sort(key=lambda row: str(row["date"]), reverse=True)
        income_total = sum(row["amount"] for row in rows if row["type"] == "Income")
        expense_total = abs(sum(row["amount"] for row in rows if row["type"] == "Expense"))
        return rows, {"income": income_total, "expenses": expense_total, "net": income_total - expense_total}

    def _compliance_expiring(self, db: Session, request: ReportRequest):
        end = request.end_date or (date.today() + timedelta(days=30))
        query = db.query(ComplianceDocument).filter(ComplianceDocument.expiry_date.isnot(None), ComplianceDocument.expiry_date <= end)
        search_filter = _contains_search(ComplianceDocument.title, ComplianceDocument.reference_number, search=request.search)
        if search_filter is not None:
            query = query.filter(search_filter)
        documents = query.order_by(ComplianceDocument.expiry_date.asc()).all()
        today = date.today()
        rows = [
            {
                "title": document.title,
                "type": _label(document.document_type),
                "reference_number": document.reference_number or "",
                "expiry_date": document.expiry_date.isoformat() if document.expiry_date else "",
                "days_left": (document.expiry_date - today).days if document.expiry_date else "",
                "status": _label(document.status),
                "responsible_person": document.responsible_person or "",
            }
            for document in documents
        ]
        return rows, {"documents": len(rows), "expired": len([row for row in rows if isinstance(row["days_left"], int) and row["days_left"] < 0])}

    def _debtors(self, db: Session, request: ReportRequest):
        """All outstanding invoice balances — the core debtors/credit report."""
        today = date.today()
        query = (
            db.query(Invoice)
            .options(joinedload(Invoice.customer))
            .filter(
                Invoice.total_amount > Invoice.paid_amount,  # outstanding balance exists
            )
        )
        if request.start_date:
            query = query.filter(Invoice.due_date >= request.start_date)
        if request.end_date:
            query = query.filter(Invoice.due_date <= request.end_date)
        if request.search:
            search_filter = _contains_search(Invoice.invoice_number, search=request.search)
            if search_filter is not None:
                query = query.filter(search_filter)
        invoices = query.order_by(Invoice.due_date.asc()).all()
        rows = []
        for inv in invoices:
            balance = max(float(inv.total_amount or 0) - float(inv.paid_amount or 0), 0)
            if balance <= 0:
                continue
            days_overdue = (today - inv.due_date).days if inv.due_date else 0
            customer_name = inv.customer.name if inv.customer else "Walk-in"
            if request.search and request.search.lower() not in customer_name.lower() and request.search.lower() not in inv.invoice_number.lower():
                continue
            rows.append(
                {
                    "customer": customer_name,
                    "invoice_number": inv.invoice_number,
                    "issue_date": inv.issue_date.isoformat() if inv.issue_date else "",
                    "due_date": inv.due_date.isoformat() if inv.due_date else "",
                    "total_amount": float(inv.total_amount or 0),
                    "paid_amount": float(inv.paid_amount or 0),
                    "balance": balance,
                    "days_overdue": max(days_overdue, 0),
                    "status": "Overdue" if days_overdue > 0 else _label(inv.status),
                }
            )
        total_outstanding = sum(row["balance"] for row in rows)
        overdue_count = len([row for row in rows if row["days_overdue"] > 0])
        return rows, {
            "total_outstanding": total_outstanding,
            "debtors": len(rows),
            "overdue": overdue_count,
        }

    def _credit_sales(self, db: Session, request: ReportRequest):
        """Credit and partial-payment sales within a date range."""
        start, end = _date_range(request)
        query = (
            db.query(Invoice)
            .options(joinedload(Invoice.customer))
            .filter(
                Invoice.issue_date >= start,
                Invoice.issue_date <= end,
                Invoice.total_amount > Invoice.paid_amount,
            )
        )
        if request.search:
            search_filter = _contains_search(Invoice.invoice_number, search=request.search)
            if search_filter is not None:
                query = query.filter(search_filter)
        invoices = query.order_by(Invoice.issue_date.desc()).all()
        rows = [
            {
                "invoice_number": inv.invoice_number,
                "date": inv.issue_date.isoformat(),
                "customer": inv.customer.name if inv.customer else "Walk-in",
                "phone": inv.customer.phone or "" if inv.customer else "",
                "total_amount": float(inv.total_amount or 0),
                "paid_amount": float(inv.paid_amount or 0),
                "balance": max(float(inv.total_amount or 0) - float(inv.paid_amount or 0), 0),
                "due_date": inv.due_date.isoformat() if inv.due_date else "",
                "status": _label(inv.status),
            }
            for inv in invoices
        ]
        return rows, {
            "total_credit": sum(row["total_amount"] for row in rows),
            "total_outstanding": sum(row["balance"] for row in rows),
            "records": len(rows),
        }

    # ── Helpers ──────────────────────────────────────────────────────

    @staticmethod
    def _get_tenant_id(db: Session) -> int | None:
        return db.info.get("tenant_id") or (db.info.get("tenant_ids") or [None])[0]

    @staticmethod
    def _parse_date(s: Any) -> date | None:
        if not s:
            return None
        try:
            return date.fromisoformat(str(s))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _int_or_none(v: Any) -> int | None:
        try:
            return int(v) if v else None
        except (ValueError, TypeError):
            return None

    # ── Accounting report handlers ────────────────────────────────────

    def _balance_sheet(self, db: Session, request: ReportRequest):
        from app.services.accounting_service import AccountingService
        extra = request.filters or {}
        tenant_id = self._get_tenant_id(db)
        svc = AccountingService(db, tenant_id=tenant_id)
        result = svc.get_balance_sheet(
            as_of_date=self._parse_date(extra.get("as_of_date")),
            branch_id=self._int_or_none(extra.get("branch_id")),
        )
        rows: list[dict[str, Any]] = []
        for section_name, section_rows in [
            ("Assets", result.get("assets", [])),
            ("Liabilities", result.get("liabilities", [])),
            ("Equity", result.get("equity", [])),
        ]:
            for r in section_rows:
                rows.append({
                    "section": section_name,
                    "account_code": r.get("account_code", ""),
                    "account_name": r.get("account_name", ""),
                    "amount": float(r.get("balance", r.get("amount", 0))),
                })
        return rows, {
            "total_assets": float(result.get("total_assets", 0)),
            "total_liabilities": float(result.get("total_liabilities", 0)),
            "total_equity": float(result.get("total_equity", 0)),
            "balanced": "YES" if result.get("is_balanced") else "NO",
        }

    def _general_ledger(self, db: Session, request: ReportRequest):
        from app.services.accounting_service import AccountingService
        extra = request.filters or {}
        account_id = self._int_or_none(extra.get("account_id"))
        if not account_id:
            return [], {}
        tenant_id = self._get_tenant_id(db)
        svc = AccountingService(db, tenant_id=tenant_id)
        result = svc.get_ledger(
            account_id=account_id,
            from_date=self._parse_date(extra.get("from_date") or request.start_date),
            to_date=self._parse_date(extra.get("end_date") or request.end_date),
            branch_id=self._int_or_none(extra.get("branch_id")),
        )
        rows = [
            {
                "date": str(e.get("date", "")),
                "reference": e.get("entry_number", ""),
                "description": e.get("description", ""),
                "debit": float(e.get("debit", 0)),
                "credit": float(e.get("credit", 0)),
                "balance": float(e.get("balance", 0)),
            }
            for e in result.get("entries", [])
        ]
        return rows, {
            "opening_balance": float(result.get("opening_balance", 0)),
            "closing_balance": float(result.get("closing_balance", 0)),
            "transactions": len(rows),
        }

    def _cashbook_report(self, db: Session, request: ReportRequest):
        from app.services.accounting_service import AccountingService
        extra = request.filters or {}
        account_id = self._int_or_none(extra.get("account_id"))
        if not account_id:
            return [], {}
        tenant_id = self._get_tenant_id(db)
        svc = AccountingService(db, tenant_id=tenant_id)
        result = svc.get_cashbook(
            account_id=account_id,
            from_date=self._parse_date(extra.get("from_date") or request.start_date),
            to_date=self._parse_date(extra.get("end_date") or request.end_date),
        )
        rows = [
            {
                "date": str(e.get("date", "")),
                "description": e.get("description", ""),
                "receipts": float(e.get("debit", 0)),
                "payments": float(e.get("credit", 0)),
                "balance": float(e.get("balance", 0)),
            }
            for e in result.get("entries", [])
        ]
        return rows, {
            "opening_balance": float(result.get("opening_balance", 0)),
            "closing_balance": float(result.get("closing_balance", 0)),
            "transactions": len(rows),
        }

    def _cash_flow_report(self, db: Session, request: ReportRequest):
        from app.services.accounting_service import AccountingService
        extra = request.filters or {}
        tenant_id = self._get_tenant_id(db)
        svc = AccountingService(db, tenant_id=tenant_id)
        result = svc.get_cash_flow(
            from_date=self._parse_date(extra.get("from_date") or request.start_date),
            to_date=self._parse_date(extra.get("end_date") or request.end_date),
            branch_id=self._int_or_none(extra.get("branch_id")),
        )
        rows = []
        for o in result.get("operating", []):
            rows.append({"section": "Operating Activities", "category": o["category"], "amount": float(o["amount"])})
        rows.append({"section": "Operating Activities", "category": "Net Cash from Operating Activities", "amount": float(result.get("total_operating", 0))})
        
        for i in result.get("investing", []):
            rows.append({"section": "Investing Activities", "category": i["category"], "amount": float(i["amount"])})
        rows.append({"section": "Investing Activities", "category": "Net Cash from Investing Activities", "amount": float(result.get("total_investing", 0))})
        
        for f in result.get("financing", []):
            rows.append({"section": "Financing Activities", "category": f["category"], "amount": float(f["amount"])})
        rows.append({"section": "Financing Activities", "category": "Net Cash from Financing Activities", "amount": float(result.get("total_financing", 0))})
        
        rows.append({"section": "Net Cash Flow", "category": "Net Increase (Decrease) in Cash", "amount": float(result.get("net_cash_flow", 0))})
        
        return rows, {
            "total_operating": float(result.get("total_operating", 0)),
            "total_investing": float(result.get("total_investing", 0)),
            "total_financing": float(result.get("total_financing", 0)),
            "net_cash_flow": float(result.get("net_cash_flow", 0)),
        }

    # ── AR / AP Aging ────────────────────────────────────────────────

    def _ar_aging(self, db: Session, request: ReportRequest):
        from collections import defaultdict
        today = date.today()
        invoices = (
            db.query(Invoice)
            .options(joinedload(Invoice.customer))
            .filter(Invoice.total_amount > Invoice.paid_amount)
            .all()
        )
        buckets: dict[int, dict[str, Any]] = defaultdict(
            lambda: {"customer": "", "current": 0.0, "days_31_60": 0.0, "days_61_90": 0.0, "over_90": 0.0, "total": 0.0}
        )
        for inv in invoices:
            balance = max(float(inv.total_amount or 0) - float(inv.paid_amount or 0), 0)
            if balance <= 0:
                continue
            key = inv.customer_id or 0
            buckets[key]["customer"] = inv.customer.name if inv.customer else "Walk-in"
            days_due = (today - inv.due_date).days if inv.due_date else 0
            if days_due <= 30:
                buckets[key]["current"] += balance
            elif days_due <= 60:
                buckets[key]["days_31_60"] += balance
            elif days_due <= 90:
                buckets[key]["days_61_90"] += balance
            else:
                buckets[key]["over_90"] += balance
            buckets[key]["total"] += balance
        rows = sorted([v for v in buckets.values() if v["total"] > 0], key=lambda r: -r["total"])
        return rows, {
            "total_outstanding": sum(r["total"] for r in rows),
            "customers": len(rows),
        }

    def _ap_aging(self, db: Session, request: ReportRequest):
        from collections import defaultdict
        today = date.today()
        # Use feed purchases as proxy for AP
        purchases = (
            db.query(FeedPurchase)
            .options(joinedload(FeedPurchase.supplier))
            .all()
        )
        buckets: dict[int, dict[str, Any]] = defaultdict(
            lambda: {"supplier": "", "current": 0.0, "days_31_60": 0.0, "days_61_90": 0.0, "over_90": 0.0, "total": 0.0}
        )
        for p in purchases:
            balance = float(p.total_amount or 0)
            if balance <= 0:
                continue
            key = p.supplier_id or 0
            buckets[key]["supplier"] = p.supplier.name if p.supplier else "Unknown"
            purchase_date = p.purchase_date or today
            days_old = (today - purchase_date).days
            if days_old <= 30:
                buckets[key]["current"] += balance
            elif days_old <= 60:
                buckets[key]["days_31_60"] += balance
            elif days_old <= 90:
                buckets[key]["days_61_90"] += balance
            else:
                buckets[key]["over_90"] += balance
            buckets[key]["total"] += balance
        rows = sorted([v for v in buckets.values() if v["total"] > 0], key=lambda r: -r["total"])
        return rows, {
            "total_payable": sum(r["total"] for r in rows),
            "suppliers": len(rows),
        }

    # ── Farm report handlers ────────────────────────────────────────

    def _cost_per_bird(self, db: Session, request: ReportRequest):
        start, end = _date_range(request, default_days=365)
        records = (
            db.query(SlaughterRecord)
            .options(joinedload(SlaughterRecord.batch))
            .filter(
                SlaughterRecord.slaughter_date >= start,
                SlaughterRecord.slaughter_date <= end,
                SlaughterRecord.total_production_cost.isnot(None),
            )
            .order_by(SlaughterRecord.slaughter_date.desc())
            .all()
        )
        rows = []
        for rec in records:
            birds = rec.live_birds_count or 1
            total = float(rec.total_production_cost or 0)
            rows.append({
                "batch": rec.batch.batch_number if rec.batch else "",
                "slaughter_date": str(rec.slaughter_date),
                "birds_in": birds,
                "dressed_weight": float(rec.total_dressed_weight or 0),
                "feed_cost": 0,
                "labour_cost": float(rec.direct_labour_cost or 0),
                "overhead_cost": float(rec.overhead_cost or 0),
                "total_cost": total,
                "cost_per_bird": round(total / max(birds, 1), 2),
                "cost_per_kg": float(rec.cost_per_kg or 0),
            })
        return rows, {"batches": len(rows), "total_cost": sum(r["total_cost"] for r in rows)}

    def _mortality_analysis(self, db: Session, request: ReportRequest):
        start, end = _date_range(request, default_days=365)
        logs = (
            db.query(MortalityLog)
            .options(joinedload(MortalityLog.batch))
            .filter(
                MortalityLog.record_date >= start,
                MortalityLog.record_date <= end,
            )
            .order_by(MortalityLog.record_date.desc())
            .all()
        )
        rows = []
        for log in logs:
            flock_size = max(log.batch.current_quantity if log.batch else 1, 1)
            rows.append({
                "date": str(log.record_date),
                "batch": log.batch.batch_number if log.batch else "",
                "cause": log.cause or "Unknown",
                "count": log.quantity,
                "pct_flock": round((log.quantity / flock_size) * 100, 2),
            })
        return rows, {"total_deaths": sum(r["count"] for r in rows), "records": len(rows)}

    def _slaughter_yield(self, db: Session, request: ReportRequest):
        start, end = _date_range(request, default_days=365)
        records = (
            db.query(SlaughterRecord)
            .options(joinedload(SlaughterRecord.batch))
            .filter(
                SlaughterRecord.slaughter_date >= start,
                SlaughterRecord.slaughter_date <= end,
            )
            .order_by(SlaughterRecord.slaughter_date.desc())
            .all()
        )
        rows = [
            {
                "batch": r.batch.batch_number if r.batch else "",
                "slaughter_date": str(r.slaughter_date),
                "birds_in": r.live_birds_count,
                "live_weight": float(r.total_live_weight or 0),
                "dressed_weight": float(r.total_dressed_weight or 0),
                "yield_pct": round(float(r.yield_percentage or 0), 2),
                "condemned": r.condemned_birds_count or 0,
                "cost_per_kg": float(r.cost_per_kg or 0),
            }
            for r in records
        ]
        return rows, {"batches": len(rows), "avg_yield": round(sum(r["yield_pct"] for r in rows) / max(len(rows), 1), 2)}

    def _branch_performance(self, db: Session, request: ReportRequest):
        from app.services.accounting_service import AccountingService
        start, end = _date_range(request, default_days=30)
        tenant_id = self._get_tenant_id(db)
        branches = (
            db.query(Branch)
            .filter(Branch.tenant_id == tenant_id, Branch.is_active.is_(True))
            .all()
        )
        svc = AccountingService(db, tenant_id=tenant_id)
        rows = []
        for branch in branches:
            try:
                pl = svc.get_profit_and_loss(
                    from_date=start,
                    to_date=end,
                    branch_id=branch.id,
                )
                revenue = float(pl.get("total_revenue") or 0)
                cogs = float(pl.get("total_cost_of_sales") or 0)
                gross = float(pl.get("gross_profit") or 0)
                expenses = float(pl.get("total_expenses") or 0)
                net = float(pl.get("net_profit") or pl.get("net_income") or 0)
                rows.append({
                    "branch": branch.name,
                    "revenue": revenue,
                    "cogs": cogs,
                    "gross_margin": gross,
                    "expenses": expenses,
                    "net_income": net,
                    "margin_pct": round(gross / max(revenue, 1) * 100, 1),
                })
            except Exception:
                pass
        return rows, {
            "total_revenue": sum(r["revenue"] for r in rows),
            "total_net": sum(r["net_income"] for r in rows),
        }

    def _csv(self, preview: ReportPreview) -> bytes:
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=preview.selected_fields)
        writer.writeheader()
        writer.writerows(preview.rows)
        return output.getvalue().encode("utf-8")

    def _xlsx(self, preview: ReportPreview) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = preview.report.title[:31]
        worksheet.append([preview.report.title])
        worksheet.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
        worksheet.append([])
        worksheet.append(preview.selected_fields)
        header_row = worksheet.max_row
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BRAND_PRIMARY.replace("#", ""))
        for row in preview.rows:
            worksheet.append([row.get(field, "") for field in preview.selected_fields])
        worksheet.freeze_panes = f"A{header_row + 1}"
        for index, field in enumerate(preview.selected_fields, start=1):
            max_width = max([len(str(field))] + [len(str(row.get(field, ""))) for row in preview.rows[:200]])
            worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_width + 2, 12), 34)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _pdf(self, preview: ReportPreview) -> bytes:
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)
        y = draw_pdf_brand_header(pdf, width, height, preview.report.title)
        left = 36
        right = width - 36
        pdf.setFillColor(colors.HexColor(BRAND_DARK))
        pdf.setFont("Helvetica", 8)
        filter_text = " | ".join(f"{key}: {value or 'Any'}" for key, value in preview.filters_applied.items())
        pdf.drawString(left, y, filter_text[:150])
        y -= 18
        pdf.drawString(left, y, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Rows: {preview.row_count}")
        y -= 22

        column_width = (right - left) / max(len(preview.selected_fields), 1)
        pdf.setFillColor(colors.HexColor(BRAND_PRIMARY))
        pdf.rect(left, y - 14, right - left, 18, fill=True, stroke=False)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 7)
        for index, field in enumerate(preview.selected_fields):
            pdf.drawString(left + (index * column_width) + 4, y - 9, field.replace("_", " ").title()[:18])
        y -= 24

        pdf.setFillColor(colors.HexColor(BRAND_DARK))
        pdf.setFont("Helvetica", 7)
        for row in preview.rows:
            if y < 44:
                pdf.showPage()
                y = draw_pdf_brand_header(pdf, width, height, preview.report.title)
            for index, field in enumerate(preview.selected_fields):
                pdf.drawString(left + (index * column_width) + 4, y, str(row.get(field, ""))[:22])
            y -= 13
        pdf.showPage()
        pdf.save()
        return buffer.getvalue()


reports_service = ReportsService()
